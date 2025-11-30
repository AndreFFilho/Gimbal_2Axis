"""
Interface gráfica em Flet do Controle do Gimbal

Este módulo monta a interface do usuário usando Flet e fornece:

- Tela principal com controles de pitch/roll (sliders) e botões de ação.
- Exibição de telemetria (pitch, roll, vbat) enviada pelo ESP32 via MQTT.
- Integração com MQTT através de callbacks de status e telemetria.
- Salvamento/abertura de "favoritos" em arquivo XLSX (openpyxl).
- Suporte a alternância de tema claro/escuro e pequenas utilidades de sistema.

Dependências principais:
- flet
- openpyxl
"""

import csv
from datetime import datetime
import os
import platform
import subprocess
import flet as ft
import multiprocessing as mp
from queue import Empty
from MQTT.mqtt_process import run_mqtt

# ===== paths =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_ARQ = os.path.join(BASE_DIR, "favoritos.xlsx")

# ===== Cabeçalho e layout da planilha =====
_CABECALHO = ["NOME", "PITCH (º)", "ROLL (º)", "HORA/DATA"]
_WIDTHS = {1: 30, 2: 14, 3: 14, 4: 30}

# ---------- utilidades ----------
def _is_wsl() -> bool:
    try:
        with open("/proc/sys/kernel/osrelease") as f:
            return "microsoft" in f.read().lower()
    except Exception:
        return False

def _abrir_arquivo(caminho: str):
    try:
        sis = platform.system()
        if _is_wsl():
            win_path = subprocess.check_output(["wslpath", "-w", caminho]).decode().strip()
            subprocess.Popen(["cmd.exe", "/C", "start", "", win_path])
        elif sis == "Windows":
            os.startfile(caminho)  # type: ignore[attr-defined]
        elif sis == "Darwin":
            subprocess.Popen(["open", caminho])
        else:
            subprocess.Popen(["xdg-open", caminho])
    except Exception as e:
        print("Erro abrindo arquivo:", e)
        pass

#Cria a planilha de setpoints favoritos
def planilha_xlsx(caminho: str) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    if not os.path.exists(caminho):
        wb = Workbook()
        ws = wb.active
        ws.title = "Favoritos"
        ws.append(_CABECALHO)
        # Larguras e congelar cabeçalho
        for idx, w in _WIDTHS.items():
            ws.column_dimensions[get_column_letter(idx)].width = w
        ws.freeze_panes = "A2"
        wb.save(caminho)
    return True

# Paleta de cores
def _theme_colors(mode: ft.ThemeMode):
    if mode == ft.ThemeMode.DARK:
        return {"bg": "#101215", "card": "#1b1f24", "green": "#2e7d32"}
    else:
        return {"bg": "#F2F4F7", "card": "#FFFFFF", "green": "#2e7d32"}

# Profundidade da interface
def depth_wrap(page: ft.Page, inner: ft.Container) -> ft.Container:
    C = _theme_colors(page.theme_mode)
    inner.bgcolor = C["card"]
    inner.border_radius = 12
    shell = ft.Container(
        bgcolor=None,
        padding=6,
        border_radius=16,
        border=None,
        content=inner,
    )
    return shell

# Estilo dos botões Inverter
def _btn_on_style(page: ft.Page) -> ft.ButtonStyle:
    C = _theme_colors(page.theme_mode)
    return ft.ButtonStyle(bgcolor={"": C["green"]}, color={"": "white"})

def _btn_off_style(page: ft.Page) -> ft.ButtonStyle:
    return ft.ButtonStyle()

# =========================== MQTT ======================
def start_mqtt_process():
    try:
        ctx = mp.get_context()  
    except ValueError:
        ctx = mp

    try:
        cmd_q = ctx.Queue()
        tel_q = ctx.Queue()
        ctl_q = ctx.Queue()
        p = ctx.Process(target=run_mqtt, args=(cmd_q, tel_q, ctl_q), daemon=True)
        p.start()
        return p, cmd_q, tel_q, ctl_q
    except Exception as e:
        print("Erro ao iniciar processo MQTT:", e)
        return None, None, None, None

def start_tel_poller(page: ft.Page, tel_q, on_telemetry_cb=None, on_status_cb=None):
    if tel_q is None:
        return None

    LOG_CSV = "gimbal_logs.csv"
    log_file = None
    log_writer = None

    try:
        file_exists = os.path.exists(LOG_CSV)
        log_file = open(LOG_CSV, "a", newline="", encoding="utf-8")
        log_writer = csv.writer(log_file)
        if not file_exists:
            # cabeçalho
            log_writer.writerow(["timestamp_pc", "level", "tag", "message"])
            log_file.flush()
    except Exception as e:
        print("Nao foi possivel abrir arquivo de log CSV:", e)
        log_file = None
        log_writer = None

    def poller():
        try:
            while True:
                try:
                    item = tel_q.get(timeout=1.0)
                except Empty:
                    continue
                except Exception:
                    break

                ttype = item.get("type")

                if ttype == "telemetry":
                    d = item.get("data", {})
                    if on_telemetry_cb:
                        try:
                            on_telemetry_cb(d.get("pitch"), d.get("roll"), d.get("vbat"))
                        except Exception:
                            pass

                elif ttype == "status":
                    if on_status_cb:
                        try:
                            on_status_cb(item.get("ok", False), item.get("msg", ""))
                        except Exception:
                            pass

                elif ttype == "error":
                    try:
                        page.snack_bar = ft.SnackBar(
                            ft.Text(f"MQTT error: {item.get('msg')}"),
                            open=True
                        )
                    except Exception:
                        pass

                elif ttype == "log" and log_writer is not None:
                    data = item.get("data", {})
                    # ESP32 manda algo como {"tag": "...", "level": "INFO", "msg": "..."}
                    level = str(data.get("level", "INFO"))
                    tag = str(data.get("tag", ""))
                    msg_txt = str(data.get("msg", ""))
                    ts = datetime.now().isoformat(timespec="seconds")

                    try:
                        log_writer.writerow([ts, level, tag, msg_txt])
                        log_file.flush()
                    except Exception:
                        pass

                page.update()
        finally:
            # fecha o arquivo de log quando a thread terminar
            if log_file is not None:
                try:
                    log_file.close()
                except Exception:
                    pass

    # inicia a thread ligada à página
    return page.run_thread(poller)


#  ===================== APP =====================
def main(pagina: ft.Page):
    # === BATERIA ===
    BAT_VMIN = 6.10   # 0% 
    BAT_VMAX = 8.20   # 100% 
    bat_warned = False  # BATERIA CRÍTICO   

    # === PAGINA ===
    pagina.title = "Controle de Gimbal"
    pagina.padding = 16
    pagina.window_min_width = 900
    pagina.window_min_height = 580
    pagina.theme_mode = ft.ThemeMode.DARK  # pode alternar pela AppBar

    # aplica 2 tons: fundo da página fixo
    C = _theme_colors(pagina.theme_mode)
    pagina.bgcolor = C["bg"]

    ok_openpyxl = planilha_xlsx(XLSX_ARQ)
    if not ok_openpyxl:
        pagina.snack_bar = ft.SnackBar(
            ft.Text("Para salvar/abrir em XLSX, instale: pip install openpyxl"),
            open=True
        )

    # ===== Status / MQTT =====
    icone_status = ft.Icon("cloud_off", size=14)
    texto_status = ft.Text("MQTT não conectado.", size=11, weight=ft.FontWeight.W_600)
    selo_status = ft.Container(
        padding=ft.padding.symmetric(6, 8),
        border_radius=16,
        content=ft.Row([icone_status, texto_status], spacing=6, alignment=ft.MainAxisAlignment.CENTER),
    )
    bat_icon = ft.Icon("battery_unknown", size=16)                 # ícone muda conforme %
    bat_txt  = ft.Text("—", size=12, weight=ft.FontWeight.W_600)   # mostrará "XX% (YY,YY V)"
    selo_bateria = ft.Container(                                  
        padding=ft.padding.symmetric(4, 6),
        border_radius=12,
        content=ft.Row([bat_icon, bat_txt], spacing=4, alignment=ft.MainAxisAlignment.CENTER),
    )

    def atualizar_status(conectado: bool, msg: str):
        texto_status.value = msg
        icone_status.name = "cloud_done" if conectado else ("sync" if ("Reconectando" in msg or "Conectando" in msg) else "cloud_off")
        pagina.update()

    # ===== Sliders =====
    titulo_cmd = ft.Text("Comandos", size=20, weight=ft.FontWeight.W_700)

    inverter_pitch = False  # começam OFF
    inverter_roll  = False

    rotulo_pitch = ft.Text("↕ Inclinação (Pitch)", size=14, weight=ft.FontWeight.W_600)
    rotulo_roll  = ft.Text("↻ Rotação (Roll)",   size=14, weight=ft.FontWeight.W_600)

    slider_pitch = ft.Slider(min=-80, max=80, divisions=160, value=0, label="{value}°")
    slider_roll  = ft.Slider(min=-80, max=80, divisions=160, value=0, label="{value}°")

    slider_pitch_wrap = ft.Container(
        content=slider_pitch,
        padding=ft.padding.only(left=6, right=6, top=8, bottom=12),
        border_radius=8,
        bgcolor=None,
        border=None,
    )
    slider_roll_wrap = ft.Container(
        content=slider_roll,
        padding=ft.padding.only(left=6, right=6, top=8, bottom=12),
        border_radius=8,
        bgcolor=None,
        border=None,
    )

    # centralizado
    valor_pitch_txt = ft.Text("0.0°", size=16, weight=ft.FontWeight.W_600, text_align=ft.TextAlign.CENTER)
    valor_roll_txt  = ft.Text("0.0°", size=16, weight=ft.FontWeight.W_600, text_align=ft.TextAlign.CENTER)

    # --- Coluna Pitch ---
    col_esq = ft.Column(
        [
            rotulo_pitch,
            slider_pitch_wrap,
            ft.Container(content=valor_pitch_txt, alignment=ft.alignment.center),
        ],
        spacing=8,
        expand=1,
    )

    # --- Coluna Roll ---
    col_dir = ft.Column( 
        [
            rotulo_roll,
            slider_roll_wrap,
            ft.Container(content=valor_roll_txt, alignment=ft.alignment.center),
        ],
        spacing=8,
        expand=1,
    )

    # --- Card Status ---
    card_cmd_inner = ft.Container(
        padding=16,
        border=None,
        content=ft.Column(
            [
                ft.Row([titulo_cmd, selo_status], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Row([col_esq, col_dir], spacing=24),
            ],
            spacing=16,
        ),
    )
    card_cmd = depth_wrap(pagina, card_cmd_inner)

    # ===== TELEMETRIA =====
    titulo_tel = ft.Text("Telemetria", size=18, weight=ft.FontWeight.W_700)
    tel_pitch_txt = ft.Text("—", size=16, text_align=ft.TextAlign.CENTER)
    tel_roll_txt  = ft.Text("—", size=16, text_align=ft.TextAlign.CENTER)

    def aplicar_telemetria(pitch_lido: float, roll_lido: float, vbat=None):  # vbat em Volts
        nonlocal bat_warned  

        tel_pitch_txt.value = f"{pitch_lido:.1f}°"
        tel_roll_txt.value  = f"{roll_lido:.1f}°"

        if vbat is not None:
            # 0% em 6,10 V e 100% em 8,20 V 
            pct = max(0, min(100, (vbat - BAT_VMIN) * (100.0 / (BAT_VMAX - BAT_VMIN))))
            bat_txt.value = f"{pct:.0f}% ({vbat:.2f} V)" 

            # Ícone Bateria Mudando
            bat_icon.name = (
                "battery_full"  if pct >= 85 else
                "battery_6_bar" if pct >= 60 else
                "battery_4_bar" if pct >= 35 else
                "battery_2_bar" if pct >= 15 else
                "battery_0_bar"
            )

            # Popup quando cair abaixo de 6,10 V
            if vbat < BAT_VMIN and not bat_warned:
                pagina.snack_bar = ft.SnackBar(ft.Text(f"Bateria baixa: {vbat:.2f} V (abaixo de {BAT_VMIN:.2f} V)"), open=True)
                bat_warned = True
            elif vbat >= BAT_VMIN:
                bat_warned = False

        pagina.update()

    card_tel_inner = ft.Container(
        padding=16,
        border=None,
        content=ft.Column(
            [
                titulo_tel,
                ft.Row(
                    [
                        ft.Column(
                            [ft.Text("↕ Inclinação (Pitch)"), ft.Container(tel_pitch_txt, alignment=ft.alignment.center)],
                            expand=1, spacing=6,
                        ),
                        ft.Column(
                            [ft.Text("↻ Rotação (Roll)"), ft.Container(tel_roll_txt, alignment=ft.alignment.center)],
                            expand=1, spacing=6,
                        ),
                    ],
                    spacing=16,
                ),
            ],
            spacing=10,
        ),
    )
    card_tel = depth_wrap(pagina, card_tel_inner)

    proc_mqtt, cmd_q, tel_q, ctl_q = start_mqtt_process()
    if proc_mqtt is None or cmd_q is None or tel_q is None or ctl_q is None:
        pagina.snack_bar = ft.SnackBar(ft.Text("Falha ao iniciar processo MQTT — funcionalidades de rede desabilitadas."), open=True)
        pagina.update()
        cmd_q = tel_q = ctl_q = None
    else:
        start_tel_poller(pagina, tel_q, on_telemetry_cb=aplicar_telemetria, on_status_cb=atualizar_status)

    # ===== BOTOES =====
    def clamp(v, vmin=-80.0, vmax=80.0):
        return max(vmin, min(vmax, v))

    def valores_comando():
        p_raw, r_raw = slider_pitch.value, slider_roll.value
        p = -p_raw if inverter_pitch else p_raw
        r = -r_raw if inverter_roll  else r_raw
        return round(clamp(p), 1), round(clamp(r), 1)

    def sincronizar_rotulos():
        p, r = valores_comando()
        valor_pitch_txt.value = f"{p:.1f}°"
        valor_roll_txt.value  = f"{r:.1f}°"

    # --- pushbuttons  ---
    def centralizar(_):
        slider_pitch.value = 0.0
        slider_roll.value  = 0.0
        ao_mover(None)
        pagina.update()

    def toggle_inverter_pitch(_):
        nonlocal inverter_pitch
        inverter_pitch = not inverter_pitch
        btn_inv_p.text = f"Inverter Inclinação ({'ON' if inverter_pitch else 'OFF'})"
        btn_inv_p.style = _btn_on_style(pagina) if inverter_pitch else _btn_off_style(pagina)
        ao_mover(None)
        pagina.update()

    def toggle_inverter_roll(_):
        nonlocal inverter_roll
        inverter_roll = not inverter_roll
        btn_inv_r.text = f"Inverter Rotação ({'ON' if inverter_roll else 'OFF'})"
        btn_inv_r.style = _btn_on_style(pagina) if inverter_roll else _btn_off_style(pagina)
        ao_mover(None)
        pagina.update()

    btn_centro = ft.ElevatedButton("Centralizar", icon="gps_fixed", on_click=centralizar)
    btn_inv_p  = ft.ElevatedButton("Inverter Inclinação (OFF)", icon="swap_vert", on_click=toggle_inverter_pitch)
    btn_inv_r  = ft.ElevatedButton("Inverter Rotação (OFF)",    icon="sync",      on_click=toggle_inverter_roll)

    bloco_botoes_inner = ft.Container(
        padding=12,
        border=None,
        content=ft.Row([btn_centro, btn_inv_p, btn_inv_r], alignment=ft.MainAxisAlignment.SPACE_EVENLY),
    )
    bloco_botoes = depth_wrap(pagina, bloco_botoes_inner)

    #=========== FAVORITOS =============
    # --- diálogos ---
    dlg_favoritar = ft.AlertDialog(modal=True)
    dlg_confirmar = ft.AlertDialog(modal=True)

    def abrir_dialog(dlg: ft.AlertDialog):
        pagina.open(dlg)

    def fechar_dialog(dlg: ft.AlertDialog):
        pagina.close(dlg)

    # --- Favoritar ---
    def favoritos_append(nome: str, p: float, r: float):
        if not ok_openpyxl:
            return

        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        wb = load_workbook(XLSX_ARQ)
        ws = wb.active
        dt = datetime.now()
        ws.append([nome, float(p), float(r), dt])
        row = ws.max_row
        ws.cell(row=row, column=2).number_format = "0.0"                     # PITCH
        ws.cell(row=row, column=3).number_format = "0.0"                     # ROLL
        ws.cell(row=row, column=4).number_format = "hh:mm:ss dd/mm/yyyy"     # HORA/DATA

        # Larguras do Excel
        for idx, w in _WIDTHS.items():
            ws.column_dimensions[get_column_letter(idx)].width = w
        wb.save(XLSX_ARQ)

    def favoritar_dialog(_):
        p, r = valores_comando()
        nome_in = ft.TextField(label="Nome do ângulo", autofocus=True)

        def salvar(_):
            nome = (nome_in.value or "").strip() or f"ÂNGULO {datetime.now().strftime('%Hh%M')}"
            try:
                favoritos_append(nome, p, r)
                fechar_dialog(dlg_favoritar)
                pagina.snack_bar = ft.SnackBar(ft.Text(f"Favorito salvo: {nome}"), open=True)
            except Exception as e:
                fechar_dialog(dlg_favoritar)
                pagina.snack_bar = ft.SnackBar(ft.Text(f"Erro ao salvar: {e}"), open=True)

        dlg_favoritar.title = ft.Text("Salvar favorito")
        dlg_favoritar.content = ft.Column(
            [ft.Text(f"↕ Inclinação: {p:.1f}°   •   ↻ Rotação: {r:.1f}°"), nome_in],
            tight=True, spacing=8
        )
        dlg_favoritar.actions = [
            ft.TextButton("Cancelar", icon="close", on_click=lambda e: fechar_dialog(dlg_favoritar)),
            ft.ElevatedButton("Salvar", icon="bookmark_add", on_click=salvar),
        ]
        dlg_favoritar.actions_alignment = ft.MainAxisAlignment.END
        abrir_dialog(dlg_favoritar)

    # --- Abrir Favrotitos ---
    def abrir_favoritos(_):
        if not ok_openpyxl:
            return
        _abrir_arquivo(XLSX_ARQ)

    # --- Excluir Favotiros ---
    def excluir_favoritos_dialog(_):
        def confirmar(_):
            try:
                if os.path.exists(XLSX_ARQ):
                    os.remove(XLSX_ARQ)
                planilha_xlsx(XLSX_ARQ)  
                fechar_dialog(dlg_confirmar)
                pagina.snack_bar = ft.SnackBar(ft.Text("Lista de favoritos excluída."), open=True)
            except Exception as e:
                fechar_dialog(dlg_confirmar)
                pagina.snack_bar = ft.SnackBar(ft.Text(f"Erro ao excluir: {e}"), open=True)

        dlg_confirmar.title = ft.Text("Confirmar exclusão")
        dlg_confirmar.content = ft.Text("Tem certeza que deseja excluir a lista de favoritos?\nEssa ação não pode ser desfeita.")
        dlg_confirmar.actions = [
            ft.TextButton("Cancelar", icon="close", on_click=lambda e: fechar_dialog(dlg_confirmar)),
            ft.ElevatedButton("Excluir", icon="delete", on_click=confirmar),
        ]
        dlg_confirmar.actions_alignment = ft.MainAxisAlignment.END
        abrir_dialog(dlg_confirmar)

    # ---  Card Favoritar ---
    card_fav_inner = ft.Container(
        padding=16,
        border=None,
        content=ft.Column(
            [
                ft.Text("Favoritos", size=18, weight=ft.FontWeight.W_700),
                ft.Row(
                    [
                        ft.ElevatedButton("Favoritar ângulo", icon="bookmark_add", on_click=favoritar_dialog),
                        ft.ElevatedButton("Abrir favoritos", icon="grid_on", on_click=abrir_favoritos),
                        ft.TextButton("Excluir lista", icon="delete", on_click=excluir_favoritos_dialog),
                    ],
                    spacing=10,
                ),
            ],
            spacing=12,
        ),
    )
    card_fav = depth_wrap(pagina, card_fav_inner)

    # ===== Publicação =====
    def ao_mover(_):
        sincronizar_rotulos()
        try:
            p, r = valores_comando()
            if cmd_q is not None:
                cmd_q.put_nowait({"pitch": float(p), "roll": float(r)})
        except Exception:
            pass

    slider_pitch.on_change = ao_mover
    slider_roll.on_change  = ao_mover

    # ===== Layout =====
    fila_inferior = ft.Row(
        [
            ft.Container(expand=1, content=card_tel),
            ft.Container(expand=1, content=card_fav),
        ],
        spacing=16,
    )

    conteudo = ft.Column(
        [
            card_cmd,
            bloco_botoes,
            fila_inferior,
        ],
        spacing=16,
        expand=True,
    )

    # ===== Alternar tema CLARO/ESCURO =====
    def alternar_tema():
        nonlocal card_cmd, card_tel, card_fav, bloco_botoes
        pagina.theme_mode = ft.ThemeMode.LIGHT if pagina.theme_mode == ft.ThemeMode.DARK else ft.ThemeMode.DARK
        pagina.bgcolor = _theme_colors(pagina.theme_mode)["bg"]
        card_cmd = depth_wrap(pagina, card_cmd_inner)
        card_tel = depth_wrap(pagina, card_tel_inner)
        bloco_botoes = depth_wrap(pagina, bloco_botoes_inner)
        card_fav = depth_wrap(pagina, card_fav_inner)
        pagina.controls.clear()
        nova_fila = ft.Row(
            [ft.Container(expand=1, content=card_tel), ft.Container(expand=1, content=card_fav)], spacing=16
        )
        novo_conteudo = ft.Column([card_cmd, bloco_botoes, nova_fila], spacing=16, expand=True)
        pagina.add(novo_conteudo)
        pagina.appbar = appbar
        pagina.update()

    # ================= Conectar GUI ao MQTT =====================
    def conectar(_):
        if ctl_q is None:
            pagina.snack_bar = ft.SnackBar(ft.Text("MQTT não iniciado — não é possível reconectar"), open=True)
            pagina.update()
            return
        try:
            ctl_q.put_nowait("reconnect")
        except Exception:
            pass

    def desconectar(_):
        try:
            ctl_q.put("shutdown")
        except Exception:
            pass

    appbar = ft.AppBar(
        title=ft.Text("Controle de Gimbal"),
        center_title=False,
        actions=[
            selo_bateria,
            ft.IconButton(icon="dark_mode", tooltip="Tema claro/escuro", on_click=lambda e: alternar_tema()),
            ft.IconButton(icon="refresh", tooltip="Reconectar", on_click=conectar),
            ft.IconButton(icon="logout", tooltip="Desconectar", on_click=desconectar),
        ],
    )

    pagina.add(conteudo)
    pagina.appbar = appbar

    # ===== FECHAR ======
    def _on_close(e):
        try:
            if ctl_q is not None:
                ctl_q.put_nowait("shutdown")
        except Exception:
            pass
        try:
            if proc_mqtt is not None:
                proc_mqtt.join(timeout=1.0)
                if proc_mqtt.is_alive():
                    proc_mqtt.terminate()
        except Exception:
            pass

    # registra handler
    try:
        pagina.on_close = _on_close
    except Exception:
        pass

    # inicialização
    def _start():
        btn_inv_p.style = _btn_off_style(pagina)
        btn_inv_r.style = _btn_off_style(pagina)
        sincronizar_rotulos()
        pagina.update()
        conectar(None)
    _start()

# === ABRIR APP ===
def construir_interface(pagina):
    return main(pagina)